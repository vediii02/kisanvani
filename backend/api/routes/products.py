# backend/api/routes/products.py

import io
import csv
import logging
from datetime import datetime, timezone
try:
    import openpyxl
except ImportError:
    openpyxl = None

from fastapi import APIRouter, Depends, HTTPException, Query, BackgroundTasks, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from typing import List, Optional
from db.session import get_db
from db.models.product import Product
from db.models.brand import Brand
from db.models.organisation import Organisation
from schemas.product import ProductResponse, ProductCreate, ProductUpdate
from core.auth import get_current_user
from db.models.user import User
from kb.loader import kb_loader

router = APIRouter(prefix="/products", tags=["products"])


@router.get("/", response_model=List[ProductResponse])
async def list_products(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    organisation_id: Optional[int] = None,
    brand_id: Optional[int] = None,
    category: Optional[str] = None,
    target_crops: Optional[str] = None,
    is_active: Optional[bool] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all products with optional filters"""
    
    query = select(Product)
    
    # Apply filters
    if organisation_id:
        query = query.filter(Product.organisation_id == organisation_id)
    
    if brand_id:
        query = query.filter(Product.brand_id == brand_id)
    
    if category:
        query = query.filter(Product.category == category)
    
    if target_crops:
        query = query.filter(Product.target_crops.contains(target_crops))
    
    if is_active is not None:
        query = query.filter(Product.is_active == is_active)
    
    query = query.offset(skip).limit(limit)
    
    result = await db.execute(query)
    products = result.scalars().all()
    
    return products


@router.get("/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get product by ID"""
    
    result = await db.execute(
        select(Product).filter(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return product


@router.get("/brand/{brand_id}", response_model=List[ProductResponse])
async def get_products_by_brand(
    brand_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all products for a brand"""
    
    result = await db.execute(
        select(Product).filter(Product.brand_id == brand_id)
    )
    products = result.scalars().all()
    
    return products


@router.get("/organisation/{org_id}", response_model=List[ProductResponse])
async def get_products_by_organisation(
    org_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all products for an organisation"""
    
    result = await db.execute(
        select(Product).filter(Product.organisation_id == org_id)
    )
    products = result.scalars().all()
    
    return products


@router.post("/", response_model=ProductResponse)
async def create_product(
    product: ProductCreate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new product"""
    
    # Verify brand exists
    result = await db.execute(
        select(Brand).filter(Brand.id == product.brand_id)
    )
    brand = result.scalar_one_or_none()
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")

    # Check for duplicate product
    existing_product_result = await db.execute(
        select(Product).where(
            and_(
                func.lower(Product.name) == product.name.lower().strip(),
                Product.company_id == product.company_id
            )
        )
    )
    if existing_product_result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="A product with this name already exists in this company")
    
    new_product = Product(**product.dict())
    db.add(new_product)
    await db.commit()
    await db.refresh(new_product)
    
    # Automatically generate embedding in the background
    background_tasks.add_task(kb_loader.load_product_to_vector_db, new_product)
    
    return new_product


@router.put("/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: int,
    product_update: ProductUpdate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update a product"""
    
    result = await db.execute(
        select(Product).filter(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    update_data = product_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(product, field, value)
    
    await db.commit()
    await db.refresh(product)
    
    # Automatically update embedding in the background
    background_tasks.add_task(kb_loader.load_product_to_vector_db, product)
    
    return product



@router.delete("/{product_id}")
async def delete_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a product"""
    
    result = await db.execute(
        select(Product).filter(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    await db.delete(product)
    await db.commit()
    
    return {"message": "Product deleted successfully"}

logger = logging.getLogger(__name__)

@router.get("/import/template/{type}")
async def get_product_import_template(
    type: str,
    current_user = Depends(get_current_user)
):
    """
    Download product import template (CSV or Excel)
    """
    headers = [
        "name", "brand_name", "category", "sub_category", "description", 
        "target_crops", "target_problems", "dosage", "usage_instructions", 
        "safety_precautions", "price_range", "is_active", "price"
    ]
    
    if type == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        # Add a sample row
        writer.writerow([
            "Sample Product", "Sample Brand", "Seeds", "Hybrid", "High yield seeds", 
            "Cotton, Wheat", "Pest attack", "2kg/acre", "Sow at 2 inch depth", 
            "Keep away from children", "500-1000", "true", "500.0"
        ])
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=products_template.csv"}
        )
    elif type == "excel":
        if not openpyxl:
            raise HTTPException(status_code=400, detail="Excel support not available.")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append([
            "Sample Product", "Sample Brand", "Seeds", "Hybrid", "High yield seeds", 
            "Cotton, Wheat", "Pest attack", "2kg/acre", "Sow at 2 inch depth", 
            "Keep away from children", "500-1000", "true", 500.0
        ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=products_template.xlsx"}
        )
    else:
        raise HTTPException(status_code=400, detail="Invalid template type. Use 'csv' or 'excel'.")

@router.post("/upload-csv")
async def upload_products_csv(
    file: UploadFile = File(...),
    company_id: int = Form(...),
    brand_id: Optional[int] = Form(None),
    brand_name: Optional[str] = Form(None),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Upload products via CSV or Excel file
    Supports providing company_id, brand_id OR brand_name (default for all rows) 
    OR per-row brand_name in the file.
    """
    is_superadmin = current_user.get("role") in ["admin", "superadmin"]
    user_org_id = current_user.get("organisation_id")
    
    logger.info(f"🔍 Product CSV Upload - File: {file.filename}, Company ID: {company_id}, Brand ID: {brand_id}, Brand Name: {brand_name}")
    
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload CSV or Excel file.")
    
    # Verify company
    from db.models.company import Company
    company_result = await db.execute(select(Company).where(Company.id == company_id))
    target_company = company_result.scalar_one_or_none()
    if not target_company:
        raise HTTPException(status_code=404, detail="Company not found")
    if not is_superadmin and target_company.organisation_id != user_org_id:
        raise HTTPException(status_code=403, detail="Access denied to this company")
    
    # Initial brand check if brand_id is provided
    default_brand = None
    if brand_id:
        brand_result = await db.execute(select(Brand).where(Brand.id == brand_id))
        default_brand = brand_result.scalar_one_or_none()
        if not default_brand:
            raise HTTPException(status_code=404, detail=f"Brand ID {brand_id} not found")
        # Check ownership
        if not is_superadmin and default_brand.organisation_id != user_org_id:
            raise HTTPException(status_code=403, detail="Access denied to this brand")
    
    # helper to get or create brand
    async def get_or_create_brand(name_to_use, org_id, co_id):
        if not name_to_use or not org_id: return None
        # Check existing
        brand_res = await db.execute(
            select(Brand).where(and_(Brand.name == name_to_use, Brand.organisation_id == org_id, Brand.company_id == co_id))
        )
        existing_brand = brand_res.scalars().first()
        if existing_brand:
            return existing_brand
        
        # Create new
        new_b = Brand(name=name_to_use, organisation_id=org_id, company_id=co_id, is_active=True)
        db.add(new_b)
        await db.flush() # Get ID without committing
        return new_b

    # Determine default brand from name if provided and ID is not
    effective_org_id = target_company.organisation_id
    if not default_brand and brand_name and effective_org_id:
        default_brand = await get_or_create_brand(brand_name, effective_org_id, company_id)

    success_count = 0
    errors = []
    brand_cache = {} # cache brand objects to avoid redundant DB hits
    if default_brand:
        brand_cache[default_brand.name] = default_brand
    
    try:
        content = await file.read()
        
        if file.filename.endswith('.csv'):
            csv_data = content.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(csv_data))
            rows = list(csv_reader)
            logger.info(f"📄 Parsed CSV - Total rows: {len(rows)}")
        else:
            if not openpyxl:
                raise HTTPException(status_code=400, detail="Excel support not available.")
            workbook = openpyxl.load_workbook(io.BytesIO(content))
            sheet = workbook.active
            file_headers = [cell.value for cell in sheet[1]]
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    rows.append(dict(zip(file_headers, row)))
            logger.info(f"📄 Parsed Excel - Total rows: {len(rows)}")
        
        if rows:
            logger.info(f"📋 CSV Headers: {list(rows[0].keys())}")
            logger.info(f"📋 First row data: {rows[0]}")
        
        for idx, row in enumerate(rows, start=2):
            try:
                if not row.get('name'):
                    logger.warning(f"⚠️ Row {idx}: Skipping - missing name. Row: {row}")
                    continue
                
                # Determine brand for this specific row
                row_brand_name = row.get('brand_name') or brand_name
                current_row_brand = default_brand
                
                if row_brand_name and effective_org_id:
                    if row_brand_name in brand_cache:
                        current_row_brand = brand_cache[row_brand_name]
                    else:
                        current_row_brand = await get_or_create_brand(row_brand_name, effective_org_id, company_id)
                        brand_cache[row_brand_name] = current_row_brand

                if not current_row_brand:
                    # Fallback to organisation level if no brand info
                    target_org_id = effective_org_id
                    target_brand_id = None
                    target_company_id = company_id
                else:
                    target_org_id = current_row_brand.organisation_id
                    target_brand_id = current_row_brand.id
                    target_company_id = company_id

                if not target_org_id:
                    error_msg = f"Row {idx}: Missing organisation context"
                    logger.error(f"❌ {error_msg}")
                    errors.append(error_msg)
                    continue

                logger.info(f"✅ Row {idx}: Creating product '{row.get('name')}' for brand {target_brand_id}")
                
                product_name = str(row.get('name', '')).strip()
                
                # Check for duplicate product
                existing_product = await db.execute(
                    select(Product).where(
                        and_(
                            func.lower(Product.name) == product_name.lower(),
                            Product.company_id == target_company_id
                        )
                    )
                )
                if existing_product.scalars().first():
                    error_msg = f"Row {idx}: Product '{product_name}' already exists in this company"
                    logger.warning(f"⏭️ {error_msg}")
                    errors.append(error_msg)
                    continue
                
                product = Product(
                    organisation_id=target_org_id,
                    company_id=target_company_id,
                    brand_id=target_brand_id,
                    name=product_name,
                    category=str(row.get('category', 'other')).strip(),
                    sub_category=str(row.get('sub_category', '')).strip() if row.get('sub_category') else None,
                    description=str(row.get('description', '')).strip() if row.get('description') else None,
                    target_crops=str(row.get('target_crops', '')).strip() if row.get('target_crops') else None,
                    target_problems=str(row.get('target_problems', '')).strip() if row.get('target_problems') else None,
                    dosage=str(row.get('dosage', '')).strip() if row.get('dosage') else None,
                    usage_instructions=str(row.get('usage_instructions', '')).strip() if row.get('usage_instructions') else None,
                    safety_precautions=str(row.get('safety_precautions', '')).strip() if row.get('safety_precautions') else None,
                    price_range=str(row.get('price_range', '')).strip() if row.get('price_range') else None,
                    price=float(row['price']) if 'price' in row and row['price'] is not None and str(row['price']).strip() != '' else None,
                    is_active=str(row.get('is_active', 'true')).lower() in ('true', '1', 'yes'),
                    created_at=datetime.now(timezone.utc)
                )
                
                db.add(product)
                await db.flush() # Get ID for vectorization
                from kb.loader import kb_loader
                background_tasks.add_task(kb_loader.load_product_to_vector_db, product)
                
                success_count += 1
                logger.info(f"✅ Row {idx}: Product '{product.name}' added to session")
                
            except Exception as e:
                error_msg = f"Row {idx}: {str(e)}"
                logger.error(f"❌ {error_msg}")
                errors.append(error_msg)
        
        logger.info(f"📊 CSV Processing Complete - Success: {success_count}, Errors: {len(errors)}")
        
        if success_count > 0:
            logger.info(f"💾 Committing {success_count} products to database...")
            await db.commit()
            logger.info(f"✅ Database commit successful!")
        else:
            logger.warning(f"⚠️ No products to commit")
        
        return {
            "message": f"CSV processed successfully. {success_count} products added to brand '{default_brand.name if default_brand else 'multiple brands'}'",
            "success_count": success_count,
            "error_count": len(errors),
            "errors": errors[:10]
        }
        
    except Exception as e:
        logger.error(f"❌ CSV Upload Failed: {str(e)}")
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to process CSV: {str(e)}")
